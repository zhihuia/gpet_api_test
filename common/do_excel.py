from openpyxl import load_workbook


class DoExcel:

    def __init__(self, filepath):
        self.wb = load_workbook(filepath)
        self.sh = self.wb["case_datas"]
        self.init_data_sh = self.wb["init_data"]

    # 读取所有测试数据
    def get_all_caseDatas(self):
        # 先获取所有的初始化数据
        init_datas = self.get_init_data()
        # 定义一个列表，专门用来存储所有的测试用例。每一个测试用例都是一个字典，有所有的相关数据
        all_casedatas = []
        for row in range(2, self.sh.max_row+1):
            new_case = {}
            new_case["method"] = self.sh.cell(row, column=5).value
            new_case["url"] = self.sh.cell(row, column=6).value
            # request_data先不赋值，拿到excel中的请求数据，用一个临时变量接收 （expected_data同理）
            temp_request_data = self.sh.cell(row, column=7).value
            temp_expected_data = self.sh.cell(row, column=8).value
            # 遍历所有的初始化值的键名，如果请求数据中，有某一个键名，则直接替换
            if temp_request_data is not None or temp_expected_data is not None:
                for item, value in init_datas.items():     # 字典的遍历
                    # 判断一下本条测试数据中，请求数据是否有需要替换的
                    if temp_request_data is not None and temp_request_data.find(item) != -1:  # !=-1说明存在，find函数
                        temp_request_data = temp_request_data.replace(item, str(value))
                    # 判断一下期望结果中，是否有要替换的数据-初始化数据
                    if temp_expected_data is not None and temp_expected_data.find(item) != -1:    # !=-1说明存在，find函数
                        temp_expected_data = temp_expected_data.replace(item, str(value))
                # 替换请求数据
                new_case["request_data"] = temp_request_data
                # 替换期望数据
                new_case["expected_data"] = temp_expected_data
                # 添加一列：匹配类型--1表示正则匹配，0表示全值匹配--测试用例中比对结果时使用
                new_case["compare_type"] = self.sh.cell(row, column=9).value
                # 添加一列：提取，如果有提取，则将提取表达式存储起来
                if self.sh.cell(row, column=10).value is not None:
                    new_case["expression"] = self.sh.cell(row, column=10).value
            all_casedatas.append(new_case)
        return all_casedatas

    def get_init_data(self):
        init_datas = {}
        # 读取所有的初始化数据，第一列为键名，第二列为键值
        for row in range(2, self.init_data_sh.max_row+1):
            key = self.init_data_sh.cell(row, column=1).value
            value = self.init_data_sh.cell(row, column=2).value
            init_datas[key] = value
        # 对初始化手机号码进行递增处理，需要2个
        phone2 = int(init_datas["${phone1}"]) + 1
        init_datas["${phone2}"] = str(phone2)
        init_datas["${phone3}"] = str(phone2 + 1)
        return init_datas

    def update_init_phone(self):
        init_phone = self.init_data_sh.cell(2, 2).value
        self.init_data_sh.cell(2, 2).value = str((int(init_phone)) + 3)

    def save_data(self, filepath):
        self.wb.save(filepath)
